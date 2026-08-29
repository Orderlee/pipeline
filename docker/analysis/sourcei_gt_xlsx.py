#!/usr/bin/env python3
"""sourcei_gt_export.py 가 낸 CSV → 눈으로 보는 용 XLSX (특이점 볼드·배경색).

호스트 anaconda 로 돈다 — openpyxl 이 컨테이너에 없고, 이 스크립트는 numpy/pandas 를 안 쓴다
(호스트 pandas 는 numpy ABI 로 깨져 있음. stdlib csv 만 쓰는 이유).

색 규약 (dataviz 상태 색과 같은 의미):
  초록 = 좋음/최고,  빨강 = 나쁨/붕괴,  주황 = 주의/특이,  회색 볼드 = 기준선
CSV 가 원자료이고 이 파일은 **보기용 사본**이다 — 숫자를 고칠 일이 있으면 CSV 쪽을 고칠 것.
"""
import csv, os, shutil, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 생성은 컨테이너(root, docker/data 안)에서 하고, **배포본은 리포 경로**에 둔다.
# docker/data 는 root 소유 + gitignore 라 사람이 열기 불편하다. 여기가 사람이 보는 자리.
GEN = "/home/user/work_p/Datapipeline-Data-data_pipeline/docker/data/fiftyone/frames_bank/report/sourcei_gt/csv"
DEST = "/home/user/work_p/Datapipeline-Data-data_pipeline/sourcei_gt_csv"
os.makedirs(DEST, exist_ok=True)
for f in sorted(glob.glob(f"{GEN}/*.csv")):
    shutil.copy2(f, DEST); os.chmod(f"{DEST}/{os.path.basename(f)}", 0o644)
SRC = DEST                      # 이후 읽기·쓰기는 전부 배포본 기준
OUT = f"{DEST}/sourcei_gt_analysis.xlsx"
GOOD = (PatternFill("solid", fgColor="C6EFCE"), Font(color="006100", bold=True))
BAD = (PatternFill("solid", fgColor="FFC7CE"), Font(color="9C0006", bold=True))
WARN = (PatternFill("solid", fgColor="FFEB9C"), Font(color="9C6500", bold=True))
BASE = (PatternFill("solid", fgColor="E7E6E6"), Font(bold=True))
HEAD = (PatternFill("solid", fgColor="44546A"), Font(color="FFFFFF", bold=True))
THIN = Border(bottom=Side("thin", color="BFBFBF"))
BASELINE_BANK, BEST_BANK, WORST_BANK = "v1.0.8.0", "v1.0.8.1", "v1.0.8.4"


def load(name):
    with open(f"{SRC}/{name}", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return rows[0], rows[1:]


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def rank_idx(vals, n, largest=True):
    """상/하위 n개의 행 인덱스 (None 제외)."""
    ok = [(i, v) for i, v in enumerate(vals) if v is not None]
    ok.sort(key=lambda x: -x[1] if largest else x[1])
    return {i for i, _ in ok[:n]}


def sheet(wb, title, name, styler=None, widths=None, numfmt="0.000"):
    hdr, rows = load(name)
    ws = wb.create_sheet(title)
    ws.append(hdr)
    for c in range(1, len(hdr) + 1):
        cell = ws.cell(1, c); cell.fill, cell.font = HEAD; cell.alignment = Alignment(wrap_text=True, vertical="center")
    for r in rows:
        ws.append([num(v) if num(v) is not None and v.strip() != "" else (v if v != "" else None) for v in r])
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{len(rows) + 1}"
    ws.row_dimensions[1].height = 42
    for c, h in enumerate(hdr, 1):
        L = get_column_letter(c)
        ws.column_dimensions[L].width = (widths or {}).get(c, min(max(10, len(h) * 0.75), 34))
        for r in range(2, len(rows) + 2):
            cell = ws.cell(r, c); cell.border = THIN
            if isinstance(cell.value, float): cell.number_format = numfmt
    if styler: styler(ws, hdr, rows)
    return ws


def paint(ws, r, c, style):
    cell = ws.cell(r + 2, c + 1); cell.fill, cell.font = style


def col_of(hdr, prefix):
    return [i for i, h in enumerate(hdr) if h.startswith(prefix)]


wb = Workbook(); wb.remove(wb.active)

# ── 00 읽는 법 ────────────────────────────────────────────────────────
idx_hdr, idx_rows = load("00_index.csv")
ws = wb.create_sheet("00_읽는법")
ws.append(["sourcei GT 검증 — 분석에 쓴 값 전부 (CSV 원자료 + 이 워크북은 보기용 사본)"])
ws["A1"].font = Font(bold=True, size=14)
ws.append([])
ws.append(["색 규약"]); ws["A3"].font = Font(bold=True, size=12)
for lab, st, desc in [("초록", GOOD, "그 열/행에서 가장 좋음 · 채택 후보"),
                      ("빨강", BAD, "붕괴·최저·잡음 의심 — 여기부터 보면 된다"),
                      ("주황", WARN, "특이/주의 — 값 자체가 이상하거나 해석에 조건이 붙음"),
                      ("회색 볼드", BASE, "기준선(v1.0.8.0) 또는 제품 현재 설정")]:
    ws.append([lab, desc]); c = ws.cell(ws.max_row, 1); c.fill, c.font = st
ws.append([])
ws.append(["먼저 볼 특이점 5가지"]); ws.cell(ws.max_row, 1).font = Font(bold=True, size=12)
for t in ["1) 04_IoU임계스윕 — 제품 임계 0.15 열(빨강)과 뱅크별 최고(초록)의 격차. 0.15 는 곡선 초입이다",
          "2) 02_클래스재현율 — 분포-IoU 의 smoke 열이 뱅크와 무관하게 통째로 빨강 (규칙 구조 문제)",
          "3) 09_frames대GT — 발화율 상위(주황)와 GT 성능 상위(초록)가 엇갈린다. 발화율은 역지표",
          "4) 06_오프셋알파 — α=1 행의 정확도 붕괴(빨강) vs α=0.3 최적(초록)",
          "5) 10_GT출처위치 — folder 출처 0.16·순번 50~99 재현율 0 (빨강) = GT 가 영상 윈도우 라벨이라는 증거"]:
    ws.append([t])
ws.append([])
ws.append(idx_hdr); r0 = ws.max_row
for c in range(1, len(idx_hdr) + 1):
    cell = ws.cell(r0, c); cell.fill, cell.font = HEAD
for r in idx_rows:
    ws.append(r)
for c, wd in zip("ABCD", (34, 70, 10, 60)):
    ws.column_dimensions[c].width = wd
for r in range(r0, ws.max_row + 1):
    ws.cell(r, 2).alignment = Alignment(wrap_text=True); ws.cell(r, 4).alignment = Alignment(wrap_text=True)


# ── 01 뱅크 × 규칙 ────────────────────────────────────────────────────
def s01(ws, hdr, rows):
    f1c = hdr.index("macro_f1_event(이벤트3클래스macroF1)")
    for rule in {r[1] for r in rows}:
        idx = [i for i, r in enumerate(rows) if r[1] == rule]
        vals = [num(rows[i][f1c]) for i in idx]
        top, bot = rank_idx(vals, 3), rank_idx(vals, 3, largest=False)
        for k, i in enumerate(idx):
            if k in top: paint(ws, i, f1c, GOOD)
            elif k in bot: paint(ws, i, f1c, BAD)
    for c in col_of(hdr, "") :
        if "_recall(" in hdr[c]:
            for i, r in enumerate(rows):
                v = num(r[c])
                if v is not None and v < 0.15: paint(ws, i, c, BAD)
    for i, r in enumerate(rows):
        if r[0] == BASELINE_BANK: paint(ws, i, 0, BASE)
        elif r[0] == BEST_BANK: paint(ws, i, 0, GOOD)
        elif r[0] == WORST_BANK: paint(ws, i, 0, WARN)


sheet(wb, "01_뱅크x규칙요약", "01_bank_rule_summary.csv", s01, widths={1: 13, 2: 17})


# ── 02 클래스 재현율 ──────────────────────────────────────────────────
def s02(ws, hdr, rows):
    for i, r in enumerate(rows):
        for c in range(1, len(hdr)):
            v = num(r[c])
            if v is None: continue
            if v < 0.15: paint(ws, i, c, BAD)
            elif v >= 0.5: paint(ws, i, c, GOOD)
        if r[0] == BASELINE_BANK: paint(ws, i, 0, BASE)


sheet(wb, "02_클래스재현율", "02_bank_class_recall.csv", s02, widths={1: 13})


# ── 03 혼동 비율 ──────────────────────────────────────────────────────
def s03(ws, hdr, rows):
    miss = [hdr.index(x) for x in hdr if "→normal" in x]
    fa = hdr.index("normal→event(정상오탐)")
    for i, r in enumerate(rows):
        for c in miss:
            v = num(r[c])
            if v is not None and v >= 0.8: paint(ws, i, c, BAD)
        v = num(r[fa])
        if v is not None and v >= 0.1: paint(ws, i, fa, WARN)
        if r[0] == BASELINE_BANK: paint(ws, i, 0, BASE)


sheet(wb, "03_혼동비율", "03_confusion_rates.csv", s03, widths={1: 13, 2: 17})


# ── 04 IoU 임계 스윕 ──────────────────────────────────────────────────
def s04(ws, hdr, rows):
    p15 = hdr.index("macro_f1_at_0.15(제품임계)"); best = hdr.index("best_macro_f1(최적macroF1)")
    beat = hdr.index("beats_topk(최적임계가topK를넘나)")
    sweep = col_of(hdr, "mf1@")
    for i, r in enumerate(rows):
        paint(ws, i, p15, BAD); paint(ws, i, best, GOOD)
        if r[beat] == "N": paint(ws, i, beat, WARN)
        vals = [num(r[c]) for c in sweep]
        bi = max(range(len(vals)), key=lambda k: (vals[k] is not None, vals[k]))
        paint(ws, i, sweep[bi], GOOD)
    c15 = sweep[[hdr[c] for c in sweep].index("mf1@0.15")]
    ws.cell(1, c15 + 1).fill, ws.cell(1, c15 + 1).font = BAD


sheet(wb, "04_IoU임계스윕", "04_iou_threshold_sweep.csv", s04, widths={1: 13})


# ── 05 홀드아웃 ───────────────────────────────────────────────────────
def s05(ws, hdr, rows):
    b = hdr.index("tuned_beats_topk"); t = hdr.index("iou_tuned(튠임계macroF1)"); p = hdr.index("iou_at_0.15(제품임계)")
    for i, r in enumerate(rows):
        paint(ws, i, b, GOOD if r[b] == "Y" else BAD)
        paint(ws, i, t, GOOD if r[b] == "Y" else WARN)
        paint(ws, i, p, BAD)


sheet(wb, "05_IoU홀드아웃", "05_iou_holdout.csv", s05, widths={1: 13})


# ── 06 오프셋 α ───────────────────────────────────────────────────────
def s06(ws, hdr, rows):
    a = hdr.index("alpha(오프셋강도)"); f1 = hdr.index("macro_f1"); ac = hdr.index("accuracy(정확도)")
    nr = hdr.index("normal_recall(normal재현율)"); bs = hdr.index("is_best_alpha(뱅크별최적)")
    for i, r in enumerate(rows):
        if r[bs] == "Y": paint(ws, i, f1, GOOD); paint(ws, i, a, GOOD)
        if num(r[a]) == 1.0: paint(ws, i, ac, BAD)
        if num(r[nr]) is not None and num(r[nr]) < 0.5: paint(ws, i, nr, BAD)
        if num(r[a]) == 0.0: paint(ws, i, 0, BASE)


sheet(wb, "06_오프셋알파", "06_offset_alpha_sweep.csv", s06, widths={1: 13})


# ── 07 마진 구간 ──────────────────────────────────────────────────────
def s07(ws, hdr, rows):
    accs = col_of(hdr, "acc[")
    tgt = hdr.index("acc[0.02~0.03]"); prev = hdr.index("acc[0.01~0.02]")
    ws.cell(1, tgt + 1).fill, ws.cell(1, tgt + 1).font = BASE
    for i, r in enumerate(rows):
        vals = [(c, num(r[c])) for c in accs if num(r[c]) is not None]
        if vals:
            paint(ws, i, min(vals, key=lambda x: x[1])[0], BAD)
        a, b = num(r[tgt]), num(r[prev])
        if a is not None and b is not None and a < b: paint(ws, i, tgt, WARN)


sheet(wb, "07_마진구간", "07_margin_bins.csv", s07, widths={1: 13})


# ── 08 카메라 × 뱅크 ──────────────────────────────────────────────────
def s08(ws, hdr, rows):
    start = hdr.index("best_bank(1위뱅크)") + 1
    for i, r in enumerate(rows):
        vals = [(c, num(r[c])) for c in range(start, len(hdr)) if num(r[c]) is not None]
        paint(ws, i, max(vals, key=lambda x: x[1])[0], GOOD)
        paint(ws, i, min(vals, key=lambda x: x[1])[0], BAD)
        paint(ws, i, start - 1, WARN)


sheet(wb, "08_카메라x뱅크", "08_camera_bank_matrix.csv", s08, widths={1: 36, 8: 14})


# ── 09 frames vs GT ───────────────────────────────────────────────────
def s09(ws, hdr, rows):
    ev = hdr.index("frames_eventrate_topk(발화율%)"); f1 = hdr.index("gt_macro_f1_topk")
    top_ev = rank_idx([num(r[ev]) for r in rows], 5)
    vals = [num(r[f1]) for r in rows]
    top, bot = rank_idx(vals, 3), rank_idx(vals, 3, largest=False)
    for i, r in enumerate(rows):
        if i in top_ev: paint(ws, i, ev, WARN)
        if i in top: paint(ws, i, f1, GOOD)
        elif i in bot: paint(ws, i, f1, BAD)
        if r[0] == BASELINE_BANK: paint(ws, i, 0, BASE)


sheet(wb, "09_frames대GT", "09_frames_vs_gt.csv", s09, widths={1: 13})


# ── 10 GT 출처·위치 ───────────────────────────────────────────────────
def s10(ws, hdr, rows):
    fold = [i for i, h in enumerate(hdr) if h.startswith("acc_folder")][0]
    cap = [i for i, h in enumerate(hdr) if h.startswith("acc_caption")][0]
    late = [i for i, h in enumerate(hdr) if h.startswith("falldown_recall[50~99]")][0]
    for i, r in enumerate(rows):
        if num(r[fold]) is not None and num(r[fold]) < 0.2: paint(ws, i, fold, BAD)
        if num(r[cap]) is not None and num(r[cap]) >= 0.9: paint(ws, i, cap, GOOD)
        v = num(r[late])
        if v is not None and v <= 0.01: paint(ws, i, late, BAD)


sheet(wb, "10_GT출처위치", "10_gt_source_and_position.csv", s10, widths={1: 13})


# ── 11 문장 원장 ──────────────────────────────────────────────────────
def s11(ws, hdr, rows):
    se, hi, tr = hdr.index("selectivity(선택도)"), hdr.index("hit(정답프레임끌어당김)"), hdr.index("trap(다른클래스가로챔)")
    for i, r in enumerate(rows):
        s, h, t = num(r[se]), num(r[hi]), num(r[tr])
        if s is None: continue
        if s >= 0.9 and h >= 50:
            paint(ws, i, se, GOOD); paint(ws, i, 1, GOOD)
        elif s < 0.3 and t >= 50:
            paint(ws, i, se, BAD); paint(ws, i, 1, BAD)
        elif t >= 100:
            paint(ws, i, tr, WARN)


sheet(wb, "11_문장원장", "11_sentences_hit_trap.csv", s11, widths={1: 11, 2: 95})


# ── 12 구문 대조 ──────────────────────────────────────────────────────
def s12(ws, hdr, rows):
    dl = hdr.index("delta(기준선대비)"); vd = hdr.index("verdict(판정)")
    for i, r in enumerate(rows):
        v = num(r[dl])
        if v is None: continue
        st = GOOD if v > 0 else BAD
        paint(ws, i, dl, st); paint(ws, i, vd, st)
        if abs(v) > 0.3: paint(ws, i, 1, st)


sheet(wb, "12_구문대조", "12_phrase_contrast.csv", s12, widths={1: 11, 2: 30})


# ── 13 템플릿 / 13b 집중도 ────────────────────────────────────────────
def s13(ws, hdr, rows):
    se = hdr.index("selectivity(선택도)")
    for cls in {r[0] for r in rows}:
        idx = [i for i, r in enumerate(rows) if r[0] == cls and num(r[se]) is not None]
        if len(idx) < 2: continue
        paint(ws, max(idx, key=lambda i: num(rows[i][se])), se, GOOD)
        paint(ws, min(idx, key=lambda i: num(rows[i][se])), se, BAD)


sheet(wb, "13_템플릿", "13_template_selectivity.csv", s13, widths={1: 11, 2: 26})


def s13b(ws, hdr, rows):
    sh = hdr.index("share_50pct(절반을내는비율%)")
    for i, r in enumerate(rows):
        if num(r[sh]) is not None and num(r[sh]) < 10: paint(ws, i, sh, WARN)


sheet(wb, "13b_문장집중도", "13b_hit_concentration.csv", s13b, widths={1: 11}, numfmt="0.0")

# ── 16/17 frames 전체 fire 조건 (SAM3 약참조) ─────────────────────────
def s16(ws, hdr, rows):
    h = hdr.index("hit_fire"); a = hdr.index("max_box_area"); d = hdr.index("fire_minus_normal")
    for i, r in enumerate(rows):
        if r[h] == "0": paint(ws, i, h, BAD)
        if num(r[a]) is not None and num(r[a]) < 0.001: paint(ws, i, a, WARN)
        if num(r[d]) is not None and num(r[d]) < 0: paint(ws, i, d, BAD)


sheet(wb, "16_fire프레임(SAM3)", "16_fire_frames_sam3.csv", s16, widths={1: 24, 17: 60, 18: 50}, numfmt="0.0000")


def s17(ws, hdr, rows):
    se, hi, tr = hdr.index("selectivity"), hdr.index("hit"), hdr.index("trap")
    for i, r in enumerate(rows):
        sv, h, t = num(r[se]), num(r[hi]), num(r[tr])
        if sv is None: continue
        if sv >= 0.8 and h >= 40: paint(ws, i, se, GOOD); paint(ws, i, 0, GOOD)
        elif sv < 0.5 and t >= 30: paint(ws, i, se, BAD); paint(ws, i, 0, BAD)


sheet(wb, "17_fire문장원장(frames)", "17_fire_sentence_ledger_frames.csv", s17, widths={1: 95, 8: 60})

def s18(ws, hdr, rows):
    rc, fp, sf = hdr.index("fire_recall"), hdr.index("fp_rate_nonfire"), hdr.index("smoke_to_fire")
    for rule in {r[1] for r in rows}:
        idx = [i for i, r in enumerate(rows) if r[1] == rule]
        for c, largest_good in ((rc, True), (fp, False), (sf, False)):
            vals = [num(rows[i][c]) for i in idx]
            top = rank_idx(vals, 3, largest=largest_good); bot = rank_idx(vals, 3, largest=not largest_good)
            for k, i in enumerate(idx):
                if k in top: paint(ws, i, c, GOOD)
                elif k in bot: paint(ws, i, c, BAD)
    for i, r in enumerate(rows):
        if r[0] == BASELINE_BANK: paint(ws, i, 0, BASE)


sheet(wb, "18_fire뱅크별(frames)", "18_fire_by_bank.csv", s18, widths={1: 13, 2: 10})

def s19(ws, hdr, rows):
    c = hdr.index("top20_specificity"); cl = hdr.index("cls")
    for cls in {r[cl] for r in rows}:
        idx = [i for i, r in enumerate(rows) if r[cl] == cls]
        vals = [num(rows[i][c]) for i in idx]
        for k, i in enumerate(idx):
            if k in rank_idx(vals, 5): paint(ws, i, c, GOOD)


sheet(wb, "19_군집x클래스부착(라벨free)", "19_cluster_class_attachment.csv", s19, widths={1: 26})


def s20(ws, hdr, rows):
    z = hdr.index("specificity_z"); cf = hdr.index("class_conflict"); cl = hdr.index("cls")
    for i, r in enumerate(rows):
        if num(r[z]) is not None and num(r[z]) >= 4: paint(ws, i, z, GOOD if r[cl] != "normal" else WARN)
        if r[cf] == "Y": paint(ws, i, cf, BAD)


sheet(wb, "20_군집별상위문장", "20_cluster_top_sentences.csv", s20, widths={1: 26, 9: 90})

def s21(ws, hdr, rows):
    lo=[i for i,h in enumerate(hdr) if h.startswith("d_mf1_ci_lo")][0]; hi=[i for i,h in enumerate(hdr) if h.startswith("d_mf1_ci_hi")][0]
    dm=[i for i,h in enumerate(hdr) if h.startswith("d_mf1_P_minus_topk")][0]
    accp=[i for i,h in enumerate(hdr) if h.startswith("hy_acc_P(")][0]; acct=[i for i,h in enumerate(hdr) if h.startswith("hy_acc_topk")][0]
    fpp=[i for i,h in enumerate(hdr) if h.startswith("fr_fp_P(")][0]
    for i, r in enumerate(rows):
        l, h = num(r[lo]), num(r[hi])
        if l is not None and l > 0: paint(ws, i, dm, GOOD)
        elif h is not None and h < 0: paint(ws, i, dm, BAD)
        if num(r[accp]) is not None and num(r[acct]) is not None and num(r[accp]) < num(r[acct]) - 0.05: paint(ws, i, accp, BAD)
        if num(r[fpp]) is not None and num(r[fpp]) > 0.15: paint(ws, i, fpp, WARN)
        if r[0] == BASELINE_BANK: paint(ws, i, 0, BASE)


sheet(wb, "21_프로토타입vs_topK", "21_prototype_vs_topk.csv", s21, widths={1: 13})


def s22c(ws, hdr, rows):
    lo=[i for i,h in enumerate(hdr) if h.startswith("ci_lo")][0]; hi=[i for i,h in enumerate(hdr) if h.startswith("ci_hi")][0]
    d=[i for i,h in enumerate(hdr) if h.startswith("d_mf1_vs_baseline")][0]; acc=[i for i,h in enumerate(hdr) if h.startswith("hy_acc")][0]
    for i, r in enumerate(rows):
        l, h = num(r[lo]), num(r[hi])
        if l is not None and l > 0: paint(ws, i, d, GOOD)
        elif h is not None and h < 0: paint(ws, i, d, BAD)
        if num(r[acc]) is not None and num(r[acc]) < 0.5: paint(ws, i, acc, BAD)


sheet(wb, "22c_부분모듈초안평가", "22c_draft_eval.csv", s22c, widths={1: 12, 3: 13})
sheet(wb, "22_부분모듈초안문장", "22_submodular_bank_draft.csv", None, widths={1: 10, 4: 90})


def s23c(ws, hdr, rows):
    es=[i for i,h in enumerate(hdr) if h.startswith("effective_share")][0]
    vals=[num(r[es]) for r in rows]
    for i, r in enumerate(rows):
        if i in rank_idx(vals, 5, largest=False): paint(ws, i, es, BAD)
        elif i in rank_idx(vals, 5): paint(ws, i, es, GOOD)


sheet(wb, "23c_허브니스뱅크별", "23c_hubness_by_bank.csv", s23c, widths={1: 13})
sheet(wb, "23_허브문장", "23_hubness.csv", None, widths={2: 12, 7: 90})
sheet(wb, "24_군집MI", "24_frame_sentence_cluster_mi.csv", None, widths={6: 60})


# ── 2026-08-27 순차 분석 큐 산출물 (§13~§18) ─────────────────────────
# 스타일러 규약: (ws, hdr, rows) 를 받고 paint(ws, 행index, 열index, 스타일) 로 칠한다.
def s33(ws, hdr, rows):
    """프루닝 — 비열등 Y 초록, 유의 개선(CI하한>0) 초록, 큰 손해 빨강, 기준선 회색."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        if row[ci["variant(프루닝안)"]] == "기준선(전량)":
            paint(ws, r, ci["bank(뱅크)"], BASE); paint(ws, r, ci["variant(프루닝안)"], BASE); continue
        lo = num(row[ci["ci_lo(2.5%)"]])
        if lo is not None and lo > 0:
            paint(ws, r, ci["d_mf1(기준선대비Δ)"], GOOD); paint(ws, r, ci["noninferior(CI하한>-0.02)"], GOOD)
        elif row[ci["noninferior(CI하한>-0.02)"]] == "Y":
            paint(ws, r, ci["noninferior(CI하한>-0.02)"], GOOD)
        elif lo is not None and lo <= -0.10:
            paint(ws, r, ci["d_mf1(기준선대비Δ)"], BAD)
        k = num(row[ci["kept_share(유지비율)"]])
        if k is not None and k <= 0.10: paint(ws, r, ci["kept_share(유지비율)"], WARN)


def s34(ws, hdr, rows):
    """방향 산술 — 개선 초록, 악화 빨강, 원본 회색."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        if row[ci["variant(문장변환)"]] == "원본":
            paint(ws, r, ci["bank(뱅크)"], BASE); paint(ws, r, ci["variant(문장변환)"], BASE); continue
        d = num(row[ci["d_mf1(원본대비Δ)"]])
        if d is not None: paint(ws, r, ci["d_mf1(원본대비Δ)"], GOOD if d > 0 else BAD)
        lo = num(row[ci["ci_lo"]])
        if lo is not None and lo > 0: paint(ws, r, ci["ci_lo"], GOOD)


def s35(ws, hdr, rows):
    """동시군집 — 특이도 최고·배경 최저 초록(큐레이션 목표), 이벤트 우세 셀 주황."""
    ci = {h: k for k, h in enumerate(hdr)}
    sd = [num(x[ci["mean_spec_sd(평균특이도SD)"]]) for x in rows]
    ms = [num(x[ci["mean_ms(평균배경)"]]) for x in rows]
    for r in rank_idx(sd, 2, True): paint(ws, r, ci["mean_spec_sd(평균특이도SD)"], GOOD)
    for r in rank_idx(ms, 2, False): paint(ws, r, ci["mean_ms(평균배경)"], GOOD)
    for r, row in enumerate(rows):
        n = num(row[ci["n_sentences(문장수)"]]) or 1
        for c in ("cls_falldown", "cls_fire", "cls_smoke"):
            v = num(row[ci[c]])
            if v is not None and v / n >= 0.25: paint(ws, r, ci[c], WARN)


def s36(ws, hdr, rows):
    """구문 β — 상위(+) 초록, 하위(−) 빨강, 부호일치 < 0.8 은 신뢰하지 말 것(주황)."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        paint(ws, r, ci["beta(부분계수)"], GOOD if row[ci["direction(방향)"]] == "상위(+)" else BAD)
        sg = num(row[ci["sign_stability(5폴드부호일치)"]])
        if sg is not None and sg < 0.8: paint(ws, r, ci["sign_stability(5폴드부호일치)"], WARN)


def s37(ws, hdr, rows):
    """계수 일치도 — ρ ≥ 0.35 초록(라벨-free 부분 대체 가능), |ρ| < 0.1 빨강(불가)."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        v = num(row[ci["spearman(ρ)"]])
        if v is None: continue
        if v >= 0.35:
            paint(ws, r, ci["spearman(ρ)"], GOOD); paint(ws, r, ci["sign_agree(부호일치율)"], GOOD)
        elif abs(v) < 0.10:
            paint(ws, r, ci["spearman(ρ)"], BAD); paint(ws, r, ci["sign_agree(부호일치율)"], BAD)


def s38(ws, hdr, rows):
    """생성 뱅크 — 최고 macro-F1 초록, 오탐 최악 2건 빨강, 공급 뱅크 회색."""
    ci = {h: k for k, h in enumerate(hdr)}
    mf = [num(x[ci["macro_f1"]]) for x in rows]
    fp = [num(x[ci["fp_normal(정상오탐)"]]) for x in rows]
    for r in rank_idx(mf, 1, True): paint(ws, r, ci["macro_f1"], GOOD)
    for r in rank_idx(fp, 2, True): paint(ws, r, ci["fp_normal(정상오탐)"], BAD)
    for r, row in enumerate(rows):
        if not row[ci["bank(뱅크)"]].startswith("GEN"): paint(ws, r, ci["bank(뱅크)"], BASE)
        lo = num(row[ci["ci_lo"]])
        if lo is not None and lo > 0:
            paint(ws, r, ci["d_mf1(v1081대비Δ)"], GOOD); paint(ws, r, ci["ci_lo"], GOOD)
        d = num(row[ci["d_mf1(v1081대비Δ)"]])
        if d is not None and d <= -0.20: paint(ws, r, ci["d_mf1(v1081대비Δ)"], BAD)


def s42(ws, hdr, rows):
    """중복 기제 — top-K 손해 빨강, argmax 손해 주황. 두 열 크기 차이가 '규칙 탓'의 몫."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r in range(len(rows)):
        paint(ws, r, ci["topk_delta"], BAD); paint(ws, r, ci["argmax_delta"], WARN)


def s44(ws, hdr, rows):
    """하향표집 대조 — 전량과 비슷하면(크기 탓 아님) 초록, GEN 열은 빨강."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        full = num(row[ci["full_iou(v1081 전량)"]]); dn = num(row[ci["down_iou(GEN크기로 하향표집)"]])
        if full is not None and dn is not None and abs(full - dn) < 0.06:
            paint(ws, r, ci["down_iou(GEN크기로 하향표집)"], GOOD)
        paint(ws, r, ci["gen_iou(GEN+pairs)"], BAD)


sheet(wb, "31_랭킹PRAUC", "31_ranking_prauc.csv", None, widths={1: 13})
sheet(wb, "32_이벤트집계", "32_event_aggregation.csv", None, widths={1: 13})
sheet(wb, "32b_이벤트홀드아웃", "32b_event_holdout.csv", None, widths={1: 13})
sheet(wb, "33_프루닝3컷", "33_pruning.csv", s33, widths={1: 12, 2: 34})
sheet(wb, "34_방향산술", "34_direction_arithmetic.csv", s34, widths={1: 12, 2: 26})
sheet(wb, "35_동시군집", "35_biclusters.csv", s35, widths={1: 10, 4: 30, 5: 34, 10: 60})
sheet(wb, "36_구문베타", "36_phrase_beta.csv", s36, widths={1: 9, 2: 11, 4: 26, 10: 11}, numfmt="0.0000")
sheet(wb, "37_베타일치도", "37_beta_agreement.csv", s37, widths={1: 11}, numfmt="0.0000")
sheet(wb, "38_생성뱅크", "38_generated_bank.csv", s38, widths={1: 36}, numfmt="0.0000")
sheet(wb, "39_생성PRAUC", "39_generated_prauc.csv", None, widths={1: 36}, numfmt="0.0000")
sheet(wb, "40_생성문장", "40_generated_sentences.csv", None, widths={1: 10, 2: 10, 3: 90}, numfmt="0.00000")
sheet(wb, "41_생성홀드아웃", "41_generated_holdout.csv", None, widths={1: 36}, numfmt="0.0000")
sheet(wb, "42_중복기제", "42_dup_mechanism.csv", s42, widths={1: 12, 15: 46}, numfmt="0.0000")
sheet(wb, "43_공정랭킹", "43_ranking_fair.csv", None, widths={1: 24}, numfmt="0.0000")
sheet(wb, "44_하향표집대조", "44_downsample_control.csv", s44, widths={1: 11}, numfmt="0.0000")


def s45(ws, hdr, rows):
    """혼합효과 — p<0.05 초록, 비유의 빨강. ICC 는 0.5 넘으면 경고(카메라가 지배)."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        pv = num(row[ci["p"]])
        if pv is not None: paint(ws, r, ci["p"], GOOD if pv < 0.05 else BAD)
        ic = num(row[ci["ICC(카메라)"]])
        if ic is not None and ic > 0.5: paint(ws, r, ci["ICC(카메라)"], WARN)


def s47(ws, hdr, rows):
    """선택법 — Δ>0 초록, Δ<-0.02 빨강. CI 가 0 을 포함하면(전부 그렇다) 판정 아님을 기억할 것."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        dv = num(row[ci["d_mf1(전량대비Δ)"]])
        if dv is None: continue
        if dv > 0.02: paint(ws, r, ci["d_mf1(전량대비Δ)"], GOOD)
        elif dv < -0.02: paint(ws, r, ci["d_mf1(전량대비Δ)"], BAD)


def s48(ws, hdr, rows):
    """AL 기준 — AUC>0.55 초록(쓸 만함), <0.5 빨강(거꾸로 간다)."""
    ci = {h: k for k, h in enumerate(hdr)}
    for r, row in enumerate(rows):
        # csv/48 은 빈 줄 뒤에 '클래스 조건부 AUC' 두 번째 헤더가 붙는 혼합 구조라
        # 뒤쪽 행은 열 수가 다르다 → 길이를 먼저 확인한다.
        j = ci.get("auc(신호→오답)")
        if j is None or j >= len(row): continue
        a = num(row[j])
        if a is None: continue
        if a > 0.55: paint(ws, r, j, GOOD)
        elif a < 0.50: paint(ws, r, j, BAD)

sheet(wb, "45_혼합효과", "45_mixed_effects.csv", s45, widths={1: 30}, numfmt="0.0000")
sheet(wb, "46_셀_뱅크x규칙x카메라", "46_cells.csv", None, widths={1: 12, 3: 26}, numfmt="0.0000")
sheet(wb, "47_DPP선택", "47_dpp_selection.csv", s47, widths={1: 12, 2: 30}, numfmt="0.0000")
sheet(wb, "48_AL기준검증", "48_al_criterion.csv", s48, widths={1: 38}, numfmt="0.0000")
sheet(wb, "49_AL후보", "49_al_candidates.csv", None, widths={2: 38, 3: 26, 5: 30, 11: 70}, numfmt="0.0000")

sheet(wb, "50_뱅크탐색", "50_optbank_search.csv", None, widths={3: 14, 5: 10}, numfmt="0.0000")
sheet(wb, "51_뱅크비교", "51_optbank_compare.csv", None, widths={1: 26}, numfmt="0.0000")
sheet(wb, "52_뱅크문장", "52_optbank_sentences.csv", None, widths={1: 10, 2: 14, 3: 90}, numfmt="0.00000")
sheet(wb, "53_규칙준수", "53_optbank_rulecheck.csv", None, widths={1: 10, 3: 20}, numfmt="0.000")
sheet(wb, "54_전버전비교", "54_optbank_vs_all.csv", None, widths={2: 16}, numfmt="0.0000")

sheet(wb, "55_전량생성_전버전", "55_genfull_vs_all.csv", None, widths={2: 22}, numfmt="0.0000")
sheet(wb, "56_전량생성_규칙", "56_genfull_rulecheck.csv", None, widths={1: 10, 3: 16}, numfmt="0.000")
sheet(wb, "57_전량생성_문장", "57_genfull_sentences.csv", None, widths={1: 10, 2: 90}, numfmt="0.00000")

wb.save(OUT)

# 폴더만 열어도 뭐가 뭔지 알게 — 파일 목록·색 규약·재생성 방법 한 장
with open(f"{DEST}/README.md", "w", encoding="utf-8") as f:
    f.write("""# sourcei GT 검증 — 분석 원자료

노션 「프롬프트 코사인 분석 GT 검증 — sourcei 7,498 프레임」 의 모든 수치.

- **`sourcei_gt_analysis.xlsx`** — 같은 값 + 강조 서식. **첫 시트 `00_읽는법`부터 보면 된다.**
  - 🟩 초록 = 그 열/행 최고·채택 후보 / 🟥 빨강 = 붕괴·최저·잡음 의심 / 🟧 주황 = 특이·주의 / ⬜ 회색 볼드 = 기준선(v1.0.8.0)
- **`*.csv`** — 원자료(utf-8-sig, Excel 에서 한글 정상). 파일별 설명은 `00_index.csv`.

## 먼저 볼 특이점
1. `04_iou_threshold_sweep` — 제품 임계 0.15 열과 뱅크별 최적(★)의 격차. 0.15 는 곡선 초입
2. `02_bank_class_recall` — 분포-IoU 의 smoke 열이 뱅크와 무관하게 전부 낮음(규칙 구조 문제)
3. `09_frames_vs_gt` — 발화율 상위와 GT 성능 상위가 엇갈림(발화율은 역지표)
4. `06_offset_alpha_sweep` — α=1 정확도 붕괴 vs α≈0.3 최적
5. `10_gt_source_and_position` — folder 출처 0.16, 낙상 순번 50~99 재현율 ≈0 (GT 가 영상 윈도우 라벨이라는 증거)
6. `16_fire_frames_sam3` / `17_fire_sentence_ledger_frames` — **frames 전체(188k)** 에서 화재 프롬프트 반응 조건(SAM3 검출 약참조, GT 아님): 점 불꽃·마진 음수에서 놓침

## 2026-08-27 순차 분석 큐 (노션 §13~§18)
7. `33_pruning` — **특이도 하위 25% 컷은 31뱅크 중 0개가 0.02 넘게 손해**(488,334→366,239 문장). 주효과 컷은 중앙값 음수. 뱅크가 클수록 더 지워도 됨(ρ=−0.449)
8. `42_dup_mechanism` — 중복컷 손해의 **65%가 규칙 탓**(top-K −0.091 vs argmax −0.032). falldown 은 문장 6개 중 5개가 근접중복이라 이 클래스 표만 사라진다
9. `34_direction_arithmetic` — **전역 문장평균 제거 후 재정규화가 31뱅크 중 30종 개선**(부호검정 p=3.0e-8). normal 중심 제거는 무효
10. `35_biclusters` — 문장군 6개 중 이벤트를 담은 건 3개뿐이고 각각 한 현장 가족에 묶임. B2(창고)가 특이도 최고·배경 최저 = 큐레이션 목표 지점
11. `38_generated_bank` — **규칙으로 만든 499문장이 12,511문장 공급 뱅크를 이긴다**(+0.034, CI[+0.007,+0.060]). 단 **정상 오탐 4.3배**. 차 벡터(GEN-diff)는 실패
12. `44_downsample_control` — 생성 뱅크의 분포-IoU 열세는 **크기 탓이 아니다**(공급 뱅크를 같은 크기로 줄여도 유지). 균질성 탓
13. `37_beta_agreement` — **라벨-free 큐레이션은 fire(ρ+0.46)·smoke(ρ+0.38)만 가능, falldown 은 ρ−0.02 로 불가**. falldown 약참조가 다음 병목
14. `36_phrase_beta` — β 상위는 장소 어휘(warehouse·escalator)다. 화이트리스트를 그대로 따르면 장소 검출기가 된다 → **이벤트 어휘 부분만** 채택

## 재생성
```
docker exec docker-analysis-1 python3 /workspace/sourcei_gt_export.py   # CSV (컨테이너)
docker exec docker-analysis-1 chmod -R 777 /data/fiftyone/frames_bank/report/sourcei_gt/csv
/home/user/anaconda3/bin/python docker/analysis/sourcei_gt_xlsx.py       # 리포로 복사 + XLSX (호스트)
```
생성은 컨테이너(`docker/data/...`, root 소유·gitignore), **사람이 보는 자리는 이 폴더**.
export 는 차트 요약값과 대조하는 assert 내장 — 표와 그림이 어긋나면 죽는다.
""")
print("저장:", OUT, os.path.getsize(OUT) // 1024, "KB, 시트", len(wb.sheetnames))
print("시트:", ", ".join(wb.sheetnames))
